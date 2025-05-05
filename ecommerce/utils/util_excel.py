import os.path

import openpyxl

"""
This method takes file name and sheet name as arguments , read the given sheet.
If the flag=Y in any of the row then fetch it 
return: a dictionary with headers as key and cell values as values 
"""
def fetch_registration_data(file_name, sheet_name):
    # Get absolute path to the directory where this script is located
    file_path = get_file_path(file_name)
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    header= [cell.value for cell in sheet[1]]
    flag_index=header.index('flag')
    # Just reading cell values use values_only=True → gives tuples like ('val1', 'val2')
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[flag_index]=='Y':
            return dict(zip(header,row))
    return None




"""
This method takes file name, sheet name, row_dict as arguments , read the given sheet.
update the flag=N for the given row
return: true if updated 
"""
def update_row_in_sheet(file_name, sheet_name, row_dict):
    try:
        file_path =get_file_path(file_name)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        header = [cell.value for cell in sheet[1]]
        sequence_index = header.index('sequence')
        flag_index = header.index('flag')
        # Modifying cell values : iter_rows() → gives Cell objects you can read/write like cell.value =
        for row in sheet.iter_rows(min_row=2):
            if row[sequence_index].value == row_dict['sequence']:
                row[flag_index].value = 'N'
                workbook.save(file_path)
                return True
        return False
    except Exception as p:
        raise RuntimeError(f'unable to update column flag! : {p}')


def get_file_path(file_name):
    base_dir = os.path.dirname(os.path.abspath(__file__))
    print(base_dir)
    # Navigate to config directory from current folder
    return os.path.join(base_dir, '..', 'configs', file_name)
