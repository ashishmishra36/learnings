import time
import pytest
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import re
from pathlib import Path


"""write a fixture which launch the browser"""
@pytest.fixture
def driver():
    # run chrome in headless mode create a object of ChromeOptions class and set the options as headless
    chrome_options= webdriver.ChromeOptions()
    # chrome_options.add_argument("headless")
    chrome_options.add_argument("--ignore-certificate-errors")
    # give this option to the driver object when we are launching the chrome
    driver = webdriver.Chrome(options=chrome_options)
    # driver = webdriver.Chrome()
    yield driver
    driver.quit()


#Tip : window.scrollBy(0,document.body.scrollHeight) - it will find the depth of the page dynamically and go to the bottom
def test_scripts(driver):
    driver.maximize_window()
    driver.get('https://rahulshettyacademy.com/AutomationPractice')
    driver.implicitly_wait(5)
    # now lets scroll to some point and take screenshot
    driver.execute_script("window.scrollBy(0,document.body.scrollHeight);")
    # so even the browser is running in headless mode screenshot can be captured
    driver.get_screenshot_as_file("screenshot.png")


def test_table_sorting(driver):
    driver.maximize_window()
    driver.get('https://rahulshettyacademy.com/seleniumPractise/#/offers')
    driver.implicitly_wait(5)
    driver.find_element(By.XPATH,"//span[text()='Veg/fruit name']").click()
    we_items= driver.find_elements(By.XPATH,"//tr/td[1]")
    items= []
    for item in we_items:
        items.append(item.text)
    print(f'items: {items}')
    # copy method will just shallow copy(not the nested list,dict) best way if you want to play with original list
    orig_list = items.copy()
    print(f'orig_list: {orig_list}')
    items.sort()
    assert items == orig_list


"""this method to do operation with excel sheet: data.xlsx"""
def test_sheet_operation():
    workbook = openpyxl.load_workbook("practice\\gfg\\data.xlsx")
    if workbook:
        sheet = workbook.sheetnames
        print(f"workbook has sheets as: {sheet}")
        if 'registration' in sheet:
            data_set= {}
            my_sheet= workbook['registration']
            print(my_sheet.cell(1,2).value)
            my_sheet.cell(3,4).value="test"
            print(my_sheet.cell(3,4).value)
            print(my_sheet.max_row, my_sheet.max_column)
            for row in my_sheet.iter_rows(min_row=1, values_only=True):
                data_set = zip(row)
            print(data_set)
        else:
            pytest.fail('could not find the sheet !')
    else:
        pytest.fail('could not find the workbook')


"""download a excel sheet and make some changes to it and then upload it """
def test_upload_download(driver):
    # first click on download
    driver.maximize_window()
    driver.get('https://rahulshettyacademy.com/upload-download-test/')
    WebDriverWait(driver,10).until(EC.presence_of_element_located((By.XPATH, "//button[@id='downloadButton']"))).click()
    print('file downloaded successfully')

    #upload a file to the
    price_ui= driver.find_element(By.XPATH, "//div[text()='Apple']/parent::div/parent::div/div[@id='cell-4-undefined']").text
    print(f'Price before the updates is : {price_ui}')
    price_sheet =None
    # update the price in the sheet
    workbook = openpyxl.load_workbook("practice\\gfg\\download.xlsx")
    if workbook:
        sheets = workbook.sheetnames
        print(sheets)
        sheet = workbook['Sheet1']
        price_sheet = sheet.cell(row=3, column=4).value
        print(price_sheet, price_ui)
        sheet.cell(row=3, column=4).value='999'
        workbook.save("practice\\gfg\\download.xlsx")


    relative_path = Path("practice\\gfg\\download.xlsx")
    absolute_path=relative_path.resolve()
    print(absolute_path)
    # upload the sheet
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//input[@type='file']"))).send_keys(str(absolute_path))
    price_after_update = driver.find_element(By.XPATH,
                                "//div[text()='Apple']/parent::div/parent::div/div[@id='cell-4-undefined']").text
    print(f'Price After the updates is : {price_after_update}')


"""click on shop button go to product page find if Blackberry product is available and select it """
def test_product_selections(driver):
    driver.maximize_window()
    driver.get("https://rahulshettyacademy.com/angularpractice/")
    driver.find_element(By.LINK_TEXT, "Shop").click()
    print(driver.current_url)
    # find all products
    products = driver.find_elements(By.XPATH,"//div[@class='card h-100']")
    # find a product matching with the given name
    for product in products:
        print(product.find_element(By.XPATH,"div/h4/a").get_attribute('text'))
        if product.find_element(By.XPATH,"div/h4/a").get_attribute('text') == 'Blackberry':
            # once produ is found click on the add button
            product.find_element(By.XPATH,"div[2]/button").click()
    # click on the checkout button
    print(driver.find_element(By.XPATH,'//*[@id="navbarResponsive"]/ul/li/a').text)
    if "Checkout ( 1 )" in driver.find_element(By.XPATH,'//*[@id="navbarResponsive"]/ul/li/a').text:
        print('Amazing We are good !!!')












