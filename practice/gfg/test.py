from multiprocessing.pool import worker

import openpyxl

arr = [4,7,8,1,2,4,1,0,9,1]
n =2
def highest(arr, n):
    # convert it to set, check id set has length> n if yes , make it sorted and find nth
    s = sorted(set(arr))
    print(s[-n])


"""
This method takes file name and sheet name as arguments , read the given sheet.
If the flag=Y in any of the row then fetch it 
return: a dictionary with headers as key and cell values as values 
"""
def fetch_registration_data(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    header= [cell.value for cell in sheet[1]]
    flag_index=header.index('flag')
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[flag_index]=='Y':
            return dict(zip(header,row))
    return None

"""
This method takes file name, sheet name, row_dict as arguments , read the given sheet.
update the flag=N for the given row
return: true if updated 
"""
def update_row_in_sheet(file_path, sheet_name, row_dict):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        header = [cell.value for cell in sheet[1]]
        sequence_index = header.index('sequence')
        flag_index = header.index('flag')
        for row in sheet.iter_rows(min_row=2):
            if row[sequence_index].value == row_dict['sequence']:
                row[flag_index].value = 'N'
                workbook.save(file_path)
                return True
        return False
    except Exception as p:
        assert False, f'unable to update column flag!'



file_path = 'C:/Users/AshishMishra/my_data/my_learning/codebase/learnings/practice/gfg/data.xlsx'
sheet_name='registration'
row_dict = fetch_registration_data(file_path, sheet_name)
update_row_in_sheet(file_path, sheet_name, row_dict)